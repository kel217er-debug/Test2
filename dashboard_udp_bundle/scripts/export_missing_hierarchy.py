#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспортирует Excel-файл с теми же колонками, что и config/employee_hierarchy.xlsx,
но содержит только сотрудников, у которых не определилась иерархия
(МРФ / Руководитель / Тимлид).

Выбор периода захардкожен ниже.
"""

import json
import os
import re
from datetime import datetime
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# -------- ПЕРИОД (править тут) --------
# Используй "current", чтобы взять meta.current_week текущую неделю (недельный режим).
# Или задай явный ключ, например "2026-W19" (неделя) или "2026-04" (месяц).
PERIOD_KEY = "2026-05"
# Один из вариантов: "week", "month", "auto"
PERIOD_KIND = "month"

# -----------------------------------

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_PROCESSED_JSON = os.path.join(PROJECT_ROOT, "data", "processed", "daily_dashboard_data.json")
DEFAULT_OUT_DIR = os.path.join(PROJECT_ROOT, "dist")


def _read_json(path: str) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _detect_period(data: Dict, period_key: str, period_kind: str) -> Tuple[str, str]:
    if period_key == "current":
        resolved = data.get("meta", {}).get("current_week")
        if not resolved:
            raise RuntimeError("meta.current_week is missing; cannot resolve PERIOD_KEY='current'")
        return resolved, "week"

    if period_kind != "auto":
        return period_key, period_kind

    # Heuristic: week keys contain 'W' (e.g. 2026-W19). Month keys look like 2026-04.
    kind = "week" if ("W" in period_key) else "month"
    return period_key, kind


def clean_name(v: object) -> str:
    s = "" if v is None else str(v).strip()
    if not s:
        return ""
    if "@" in s:
        parts = s.split()
        kept: List[str] = []
        for p in parts:
            if "@" in p:
                break
            kept.append(p)
        s = " ".join(kept)
    tokens = re.findall(r"[A-Za-zА-Яа-яЁё-]+", s)
    return " ".join(tokens).strip()


def _is_missing(info: Dict) -> bool:
    # "Не определилось" трактуем как пустые значения.
    return (not info.get("mrf")) or (not info.get("director")) or (not info.get("teamlead"))


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def export_missing(processed_json_path: str, out_path: str) -> int:
    data = _read_json(processed_json_path)
    period_key, period_kind = _detect_period(data, PERIOD_KEY, PERIOD_KIND)

    # Берём сотрудников так же, как вкладка "Сотрудники" в HTML:
    # DATA.employees, но только e.in_data === true.
    missing: List[Dict] = []
    seen: Set[str] = set()
    for e in data.get("employees", []):
        if not isinstance(e, dict):
            continue
        if not e.get("in_data"):
            continue
        name = clean_name(e.get("name", ""))
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        row = dict(e)
        row["name"] = name
        if _is_missing(row):
            missing.append(row)

    missing.sort(key=lambda x: x.get("name", ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing"

    header = ["МРФ", "Направление", "ФИО", "Тимлидер", "Руководитель", "Действующий да/нет"]
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for e in missing:
        ws.append(
            [
                e.get("mrf", ""),
                e.get("direction", ""),
                e.get("name", ""),
                e.get("teamlead", ""),
                e.get("director", ""),
                "да" if bool(e.get("is_active", True)) else "нет",
            ]
        )

    # Metadata rows (below the table, so users can filter table freely)
    ws["A" + str(len(missing) + 3)] = "Период"
    ws["B" + str(len(missing) + 3)] = f"{period_key} ({period_kind})"
    ws["A" + str(len(missing) + 4)] = "Источник"
    ws["B" + str(len(missing) + 4)] = os.path.relpath(processed_json_path, PROJECT_ROOT)
    ws["A" + str(len(missing) + 5)] = "Строк"
    ws["B" + str(len(missing) + 5)] = len(missing)
    for r in range(len(missing) + 3, len(missing) + 6):
        ws[f"A{r}"].font = Font(bold=True)

    _autosize(ws)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    saved_path = out_path
    try:
        wb.save(saved_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        saved_path = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        wb.save(saved_path)
        print(f"WARN: could not overwrite (file may be open). Saved to: {os.path.abspath(saved_path)}")
    else:
        print(f"Saved: {os.path.abspath(saved_path)}")
    print(f"Period: {period_key} ({period_kind})")
    print(f"Rows: {len(missing)}")
    return 0


def main() -> int:
    processed = DEFAULT_PROCESSED_JSON
    period_key, period_kind = _detect_period(_read_json(processed), PERIOD_KEY, PERIOD_KIND)
    out_name = f"missing_employee_hierarchy_{period_key}_{period_kind}.xlsx".replace(":", "_")
    out_path = os.path.join(DEFAULT_OUT_DIR, out_name)
    return export_missing(processed, out_path)


if __name__ == "__main__":
    raise SystemExit(main())
