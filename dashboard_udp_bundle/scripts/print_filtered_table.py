#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Builds a filtered employees Excel table from processed dashboard JSON."""

import argparse
import json
import os
import re
from typing import Dict, List, Tuple

from python_calamine import CalamineWorkbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from dashboard_logic.excel_filters import COL, is_our_channel
from dashboard_logic.periods import iso_week_key, month_key

# In processed JSON, primary channels are represented by directions:
# NOD -> NIR, Partners NOD -> NPP, NPGS -> NSP.
ALLOWED_DIRECTIONS = {"НИР", "НПП", "НСП"}
_DT_RE = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})\s*$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Print filtered table from data/processed/daily_dashboard_data.json"
    )
    parser.add_argument(
        "--data",
        default=os.path.join("data", "processed", "daily_dashboard_data.json"),
        help="Path to processed JSON file",
    )
    parser.add_argument("--direction", default="", help="Filter by direction")
    parser.add_argument("--mrf", default="", help="Filter by MRF")
    parser.add_argument("--director", default="", help="Filter by director")
    parser.add_argument("--teamlead", default="", help="Filter by teamlead")
    parser.add_argument(
        "--period",
        default="current",
        help="Period key (e.g. 2026-W18 or 2026-04). Use 'current' for meta.current_week.",
    )
    parser.add_argument(
        "--kind",
        choices=("week", "month", "auto"),
        default="auto",
        help="Period kind. 'auto' detects from period key format.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limit number of output rows (0 = no limit)",
    )
    parser.add_argument(
        "--out",
        default=os.path.join("dist", "filtered_table.xlsx"),
        help="Output XLSX path",
    )
    parser.add_argument(
        "--export-included-rows",
        action="store_true",
        help="Export source Excel rows that project includes into calculations",
    )
    parser.add_argument(
        "--source",
        default=os.path.join("data", "source", "muz.xlsx"),
        help="Path to source MUZ Excel file (.xlsx)",
    )
    return parser.parse_args()


def read_json(path: str) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_dt(v):
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        m = _DT_RE.match(s)
        if not m:
            return None
        d, mo, y, h, mi, sec = map(int, m.groups())
        from datetime import datetime

        try:
            return datetime(y, mo, d, h, mi, sec)
        except ValueError:
            return None
    return None


def detect_period(data: Dict, period_arg: str, kind_arg: str) -> Tuple[str, str]:
    if period_arg == "current":
        period_key = data.get("meta", {}).get("current_week")
        period_kind = "week"
        return period_key, period_kind

    period_key = period_arg
    if kind_arg != "auto":
        return period_key, kind_arg

    period_kind = "month" if "-" in period_key and "W" not in period_key else "week"
    return period_key, period_kind


def filter_employees(data: Dict, direction: str, mrf: str, director: str, teamlead: str) -> List[Dict]:
    out = []
    for e in data.get("employees", []):
        if not e.get("in_data"):
            continue
        # Keep only applications from NOD / Partners NOD / NPGS scopes.
        if e.get("direction") not in ALLOWED_DIRECTIONS:
            continue
        if direction and e.get("direction") != direction:
            continue
        if mrf and e.get("mrf") != mrf:
            continue
        if director and e.get("director") != director:
            continue
        if teamlead and e.get("teamlead") != teamlead:
            continue
        out.append(e)
    return out


def get_bucket(data: Dict, period_key: str, period_kind: str, agg: str, level: str) -> Dict:
    kind_key = "by_month" if period_kind == "month" else "by_week"
    return data.get(agg, {}).get(kind_key, {}).get(period_key, {}).get(level, {})


def fmt_pct(v: float) -> str:
    return f"{v:.2f}%"


def fmt_money(v: float) -> str:
    return f"{v:,.2f}".replace(",", " ")


def make_rows(data: Dict, emps: List[Dict], period_key: str, period_kind: str) -> List[Dict]:
    reg_emp = get_bucket(data, period_key, period_kind, "registered", "employee")
    cls_emp = get_bucket(data, period_key, period_kind, "closed", "employee")
    open_emp = data.get("open", {}).get("employee", {})

    rows = []
    for e in emps:
        name = e.get("name", "")
        r = reg_emp.get(name, {})
        c = cls_emp.get(name, {})
        o = open_emp.get(name, {})
        row = {
            "name": name,
            "teamlead": e.get("teamlead", "") or "—",
            "director": e.get("director", "") or "—",
            "direction": e.get("direction", "") or "—",
            "mrf": e.get("mrf", "") or "—",
            "n_total": r.get("n_total", 0),
            "sla_acc_rate": r.get("sla_acc_rate", 0.0),
            "transfer_rate": r.get("transfer_rate", 0.0),
            "n_connected": c.get("n_connected", 0),
            "nd_sum": c.get("nd_sum", 0.0),
            "equip_share": c.get("equip_share", 0.0),
            "crossell_pct": c.get("crossell_pct", 0.0),
            "fraud_pct": c.get("fraud_pct", 0.0),
            "n_open": o.get("n_open", 0),
            "n_stale30": o.get("n_stale30", 0),
        }
        if row["n_total"] > 0 or row["n_connected"] > 0 or row["n_open"] > 0:
            rows.append(row)
    rows.sort(key=lambda x: x["n_total"], reverse=True)
    return rows


def write_excel(rows: List[Dict], period_key: str, period_kind: str, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "filtered_table"

    ws["A1"] = "Period"
    ws["B1"] = f"{period_key} ({period_kind})"
    ws["A2"] = "Rows"
    ws["B2"] = len(rows)
    ws["A1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)

    start_row = 4
    headers = [
        "name",
        "teamlead",
        "director",
        "direction",
        "mrf",
        "n_total",
        "sla_acc_rate",
        "transfer_rate",
        "n_connected",
        "nd_sum",
        "equip_share",
        "crossell_pct",
        "fraud_pct",
        "n_open",
        "n_stale30",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = start_row + 1
    for row in rows:
        ws.cell(row=r, column=1, value=row["name"])
        ws.cell(row=r, column=2, value=row["teamlead"])
        ws.cell(row=r, column=3, value=row["director"])
        ws.cell(row=r, column=4, value=row["direction"])
        ws.cell(row=r, column=5, value=row["mrf"])
        ws.cell(row=r, column=6, value=row["n_total"])
        ws.cell(row=r, column=7, value=float(row["sla_acc_rate"]) / 100.0)
        ws.cell(row=r, column=8, value=float(row["transfer_rate"]) / 100.0)
        ws.cell(row=r, column=9, value=row["n_connected"])
        ws.cell(row=r, column=10, value=float(row["nd_sum"]))
        ws.cell(row=r, column=11, value=float(row["equip_share"]) / 100.0)
        ws.cell(row=r, column=12, value=float(row["crossell_pct"]) / 100.0)
        ws.cell(row=r, column=13, value=float(row["fraud_pct"]) / 100.0)
        ws.cell(row=r, column=14, value=row["n_open"])
        ws.cell(row=r, column=15, value=row["n_stale30"])
        r += 1

    data_last_row = max(start_row + 1, r - 1)
    _apply_formats(ws, start_row + 1, data_last_row)
    _autosize(ws)

    out_abs = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_abs), exist_ok=True)
    wb.save(out_abs)
    print(f"XLSX saved: {out_abs}")


def _apply_formats(ws: Worksheet, data_start: int, data_end: int) -> None:
    pct_cols = [7, 8, 11, 12, 13]
    for col in pct_cols:
        for row in range(data_start, data_end + 1):
            ws.cell(row=row, column=col).number_format = "0.00%"

    for row in range(data_start, data_end + 1):
        ws.cell(row=row, column=10).number_format = "# ##0.00"

    for row in range(data_start, data_end + 1):
        for col in range(6, 16):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center")

def _autosize(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def _find_source_file(path_arg: str) -> str:
    if os.path.exists(path_arg):
        return path_arg
    source_dir = os.path.join("data", "source")
    if not os.path.isdir(source_dir):
        return path_arg
    for name in os.listdir(source_dir):
        lower_name = name.lower()
        if lower_name.endswith(".xlsx") and ("muz" in lower_name or "муз" in lower_name):
            return os.path.join(source_dir, name)
    return path_arg


def export_included_rows(source_path: str, period_key: str, period_kind: str, out_path: str) -> None:
    source_path = _find_source_file(source_path)
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Source file not found: {source_path}")

    wb_src = CalamineWorkbook.from_path(source_path)
    sheet = wb_src.get_sheet_by_name("Sheet1")
    rows = list(sheet.to_python())
    if not rows:
        raise RuntimeError("Source sheet is empty")

    header = rows[0]
    included = []
    for row in rows[1:]:
        try:
            primary_podr = str(row[COL["primary_podr"]]).strip() if row[COL["primary_podr"]] is not None else ""
        except IndexError:
            continue
        if not is_our_channel(primary_podr):
            continue

        try:
            reg_dt = parse_dt(row[COL["reg_dt"]])
        except IndexError:
            reg_dt = None
        if reg_dt is None:
            continue

        if period_kind == "week":
            key, _, _ = iso_week_key(reg_dt)
        else:
            key = month_key(reg_dt)
        if key != period_key:
            continue
        included.append(row)

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "included_rows"

    ws["A1"] = "Source file"
    ws["B1"] = os.path.abspath(source_path)
    ws["A2"] = "Period"
    ws["B2"] = f"{period_key} ({period_kind})"
    ws["A3"] = "Included rows"
    ws["B3"] = len(included)
    ws["A1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    start_row = 5
    for col, name in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=col, value=str(name) if name is not None else "")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = start_row + 1
    for row in included:
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1

    _autosize(ws)
    out_abs = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_abs), exist_ok=True)
    wb_out.save(out_abs)
    print(f"XLSX saved: {out_abs}")


def main() -> int:
    args = parse_args()
    data = read_json(args.data)

    period_key, period_kind = detect_period(data, args.period, args.kind)
    if not period_key:
        print("ERROR: period is not resolved.")
        return 1

    if args.export_included_rows:
        export_included_rows(
            source_path=args.source,
            period_key=period_key,
            period_kind=period_kind,
            out_path=args.out,
        )
        return 0

    emps = filter_employees(data=data, direction=args.direction, mrf=args.mrf, director=args.director, teamlead=args.teamlead)
    rows = make_rows(data, emps, period_key, period_kind)

    if args.limit > 0:
        rows = rows[: args.limit]

    write_excel(rows, period_key, period_kind, args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#  команда запуска создание таблици с фильтрами & "C:\Users\weren\AppData\Local\Programs\Python\Python312\python.exe" scripts\print_filtered_table.py --out dist\filtered_table.xls
